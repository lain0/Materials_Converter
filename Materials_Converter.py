#  Программа разбирает xml файл "materials.xml", в котором
#  содержится информация о режимах обработки станка лазерной резки
#  семейства Unimach и выводит параметры в Excel таблицу "Materials(X)кВт",
#  где X-выбранная мощность излучателя, которую надо задать в переменной Power.
#  Каждому режиму обработки соответствует отдельный лист таблицы. Имя листа
#  таблицы это ID режима в xml файле. Для навигации на первом листе таблицы
#  помещено оглавление с гиперссылками на имена режимов, как они указаны
#  в станке. Листы таблицы сразу могут быть распечатаны,
#  т.к. все настройки печати уже заданы.

import os
import openpyxl
import lxml.etree

from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Alignment

#  Путь к файлу materials.xml
#  Path = r'd:\петербург\materials.xml'
Path = 'materials.xml'

TypeOfMaterial = [
                 'Сталь', 'Нерж.', 'Оцинк.', 'Ал.', 'Медь', 'Юстировка',
                 'Бронза', 'Латунь', 'Текстолит', 'Паронит', 'Выпаривание',
                 'Гравировка'
                 ]

PowerType = [
            '300 Вт', '500 Вт', '750 Вт', '1 КВт', '1.5 КВт', '2 КВт', '3 КВт',
            '4 КВт', '5 КВт', '6 КВт', '10 КВт', '15 КВт', '20 КВт', '25 КВт'
            ]

TypeOfRegulation = ['Без регулирования', 'По амплитуде', 'По длительности']
TypeOfGas = ["Кислород", "Воздух", "Углекислота", "Азот"]

print([str(PowerType[i])+' -  '+str(i+1) for i in range(0, len(PowerType)-1)])
print('Выберите нужную мощность')

Power = input()
#  Задание нужной мощности излучателя
#  Power = '7'
YesNot = ["Нет", "	Да"]
A = {}
#  Создание новой книги Excel
Materials = Workbook()
Materials['Sheet'].column_dimensions['A'].width = 25


#  Функция оформления листа режима в Excel
def NamesAndSpaces():

    #  Стили заливок для ячеек

    My_FillBlue = PatternFill(
        start_color='538ed5', end_color='538ed5', fill_type='solid')
    my_fill = PatternFill(
        start_color='ccffcc', end_color='ccffcc', fill_type='solid')
    my_fillred = PatternFill(
        start_color='ed5a69', end_color='ed5a69', fill_type='solid')

    #  Ширина столбцов
    Material.column_dimensions['A'].width = 22
    Material.column_dimensions['B'].width = 25
    Material.column_dimensions['C'].width = 17
    Material.column_dimensions['D'].width = 10
    Material.column_dimensions['E'].width = 12
    Material.column_dimensions['F'].width = 25
    Material.column_dimensions['G'].width = 17
    Material.column_dimensions['H'].width = 10

    Material['A1'].value = 'Параметры обработки'
    Material['A1'].fill = my_fillred
    Material['A35'].value = 'Газовая консоль'
    Material['A35'].fill = my_fillred
    Material['E1'].value = 'Излучатель'
    Material['E1'].fill = my_fillred
    Material['E24'].value = 'Z лазер'
    Material['E24'].fill = my_fillred
    Material['E26'].value = 'Резка'
    Material['E24'].value = 'Z лазер'
    Material['E35'].value = 'Выпаривание'
    Material['E35'].fill = my_fillred

    Green = [
            "B2", "B8", "B12", "B15", "B18", "B22", "B23", "B24", "B25",
            "B26", "B27", "B28", "B29", "B30", "B32", "B33", "B36", "B39",
            "F2", "F6", "F10", "F14", "F18", "F22", "F23", "F25", "F27",
            "F28", "F29", "F30", "F31", "F32", "F33", "F35", "F37", "F38",
            "F39", "F40", "F41", "F42", "F43"
            ]
    #  Заливка зеленым
    for i in range(0, len(Green)):
        Material[Green[i]].fill = my_fill
    #  Задание имен заголовкам
    Material['B2'].value = 'Резка'
    Material['B8'].value = 'Положение сопла'
    Material['B12'].value = 'Способ входа'
    Material['B15'].value = 'Способ выхода'
    Material['B18'].value = 'Расставлять перемычки'
    Material['B22'].value = 'Автоочистка сопла'
    Material['B23'].value = 'Врезок до чистки'
    Material['B24'].value = 'Обдув при пробивке'
    Material['B25'].value = 'Охлаждение при резке'
    Material['B26'].value = 'Автокалибровка'
    Material['B27'].value = 'Предварительный прожиг'
    Material['B28'].value = 'Скорость гравировки '
    Material['B29'].value = 'Скорость выпаривания'
    Material['B30'].value = 'Включить выпаривание'
    Material['B32'].value = 'Линза'
    Material['B33'].value = 'Сопло'
    Material['B36'].value = 'Резка'
    Material['B39'].value = 'Выпаривание'
    Material['F2'].value = 'Резка'
    Material['F6'].value = 'Прожиг'
    Material['F10'].value = 'Гравировка'
    Material['F14'].value = 'FineCut'
    Material['F18'].value = 'Выпаривание'
    Material['F22'].value = 'Мин.  уровень мощности'
    Material['F23'].value = 'Способ регулирования'
    Material['F25'].value = 'Без слежения'
    Material['F27'].value = 'Высота подскока'
    Material['F28'].value = 'Время опускания'
    Material['F29'].value = 'Подскок линзы'
    Material['F30'].value = 'Высота фокуса'
    Material['F31'].value = 'Время подскока'
    Material['F32'].value = 'Положение линзы'
    Material['F33'].value = 'Высота гравировки'
    Material['F35'].value = 'Без слежения'
    Material['F37'].value = 'Высота подскока'
    Material['F38'].value = 'Время опускания'
    Material['F39'].value = 'Подскок линзы'
    Material['F40'].value = 'Высота фокуса'
    Material['F41'].value = 'Время подскока'
    Material['F42'].value = 'Положение линзы'
    Material['F43'].value = 'Высота гравировки'
    Material['C3'].value = 'Скорость'
    Material['C4'].value = 'FineCut'
    Material['C5'].value = 'Время прожига'
    Material['C6'].value = 'Эквидистанта'
    Material['C13'].value = 'Радиус'
    Material['C14'].value = 'Длина'
    Material['C16'].value = 'Длина/Радиус'
    Material['C19'].value = 'Длина перемычки'
    Material['C20'].value = 'Промежуток'
    Material['C21'].value = 'Мин. длина элем.'
    Material['C23'].value = 'Кол-во врезок'
    Material['C37'].value = 'Газ'
    Material['C38'].value = 'Давление(бар)'
    Material['C40'].value = 'Газ'
    Material['C41'].value = 'Давление(бар)'
    Material['G3'].value = 'Амплитуда(%)'
    Material['G4'].value = 'Частота(Гц)'
    Material['G5'].value = 'Заполнение(%)'
    Material['G7'].value = 'Амплитуда(%)'
    Material['G8'].value = 'Частота(Гц)'
    Material['G9'].value = 'Заполнение(%)'
    Material['G11'].value = 'Амплитуда(%)'
    Material['G12'].value = 'Частота(Гц)'
    Material['G13'].value = 'Заполнение(%)'
    Material['G15'].value = 'Амплитуда(%)'
    Material['G16'].value = 'Частота(Гц)'
    Material['G17'].value = 'Заполнение(%)'
    Material['G19'].value = 'Амплитуда(%)'
    Material['G20'].value = 'Частота(Гц)'
    Material['G21'].value = 'Заполнение(%)'
    Material['G26'].value = 'Положение'
    Material['G36'].value = 'Положение'

    #  Настройка области печати и параметров листа
    Material.print_area = 'A1:H43'
    Material.page_setup.orientation = 'landscape'
    Material.page_setup.paperSize = Material.PAPERSIZE_A4
    Material.page_setup.scale = 70
    Material.page_setup.fitToHeight = 1
    Material.page_setup.fitToWidth = 1
    return


#  Парсинг xml файла
xml = lxml.etree.parse(Path)
root = xml.getroot()
#  Счетчик для индексации ячеек оглавления
i = 1
#  Итерация по узлам xml файла с тегами Material и мощностью Power
for element in root.iter("Material"):
    if element.get('Power') == Power:

        #  Тип режима, определяет материал
        Type = int(element.get('Type'))
        #  Создается страница книги Excel. Название книги-ID режима
        Materials.create_sheet(element.get('ID'))
        Material = Materials[element.get('ID')]

#  В ячейку B1 заносится название режима, как оно указано в станке
#  Ячейки B1,C1,D1 объединяются, текст выравнивается по центру
        Material.merge_cells('B1:D1')
        Material['B1'].value = (
            '   ' +
            (TypeOfMaterial[int(Type)] + ' ' +
                str(float(element.get('Thickness'))) +
                ' мм. ' + element.get('Description')))

        Material['B1'].font = Font(size=24)
        Material['B1'].alignment = Alignment(horizontal='center')
        #  Вызов функции оформления страницы
        NamesAndSpaces()

#  Zлазер
        #  Высота фокуса
        Material['G30'].value = int(
            element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "FocusPosition"]').get('Value'))/10

        #  Высота подскока
        Material['G27'].value = int(
            element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "JumpHeight"]').get('Value'))/10

        #  Время подскока
        Material['G31'].value = int(
            element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "JumpTime"]').get('Value'))

        #  Положение линзы
        Material['G32'].value = int(
            element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "LensPosition"]').get('Value'))/10

        #  Подскок линзы
        Material['G29'].value = int(
            element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "LensJump"]').get('Value'))/10

        #  Высота гравировки
        Material['G33'].value = int(
            element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "EngraveHeight"]').get('Value'))/10

        #  Время опускания
        Material['G28'].value = int(
            element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "SlowDownTime"]').get('Value'))
        try:
            #  Без слежения
            Material['G25'].value = YesNot[int(
                int(element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "WithoutTracking"]').get('Value')))]
            #  Положение
            Material['H26'].value = int(
                element.find('Parameter/[@Name = "THC_THC"]/\
ParamValue[@Name = "ManualPosition"]').get('Value'))
        except AttributeError:
            pass


#  Излучатель
        #  Резка
        Material['H3'].value = float(
            element.find('Parameter/[@Name = "FLC0"]/\
ParamValue[@Name = "Amplitude"]').get('Value'))

        Material['H4'].value = float(
            element.find('Parameter/[@Name = "FLC0"]/\
ParamValue[@Name = "Frequency"]').get('Value'))

        Material['H5'].value = float(
            element.find('Parameter/[@Name = "FLC0"]/\
ParamValue[@Name = "PulseLength"]').get('Value'))

        #  Прожиг
        Material['H7'].value = float(
            element.find('Parameter/[@Name = "FLC1"]/\
ParamValue[@Name = "Amplitude"]').get('Value'))

        Material['H8'].value = float(
            element.find('Parameter/[@Name = "FLC1"]/\
ParamValue[@Name = "Frequency"]').get('Value'))

        Material['H9'].value = float(
            element.find('Parameter/[@Name = "FLC1"]/\
ParamValue[@Name = "PulseLength"]').get('Value'))

        #  Гравировка
        Material['H11'].value = float(
            element.find('Parameter/[@Name = "FLC2"]/\
ParamValue[@Name = "Amplitude"]').get('Value'))

        Material['H12'].value = float(
            element.find('Parameter/[@Name = "FLC2"]/\
ParamValue[@Name = "Frequency"]').get('Value'))

        Material['H13'].value = float(
            element.find('Parameter/[@Name = "FLC2"]/\
ParamValue[@Name = "PulseLength"]').get('Value'))

        #  FineCut
        Material['H15'].value = float(
            element.find('Parameter/[@Name = "FLC3"]/\
ParamValue[@Name = "Amplitude"]').get('Value'))

        Material['H16'].value = float(
            element.find('Parameter/[@Name = "FLC3"]/\
ParamValue[@Name = "Frequency"]').get('Value'))

        Material['H17'].value = float(
            element.find('Parameter/[@Name = "FLC3"]/\
ParamValue[@Name = "PulseLength"]').get('Value'))

        #  Выпаривание(может быть не прописано в режиме)
        try:
            Material['H19'].value = float(
                element.find('Parameter/[@Name = "FLC4"]/\
ParamValue[@Name = "Amplitude"]').get('Value'))

            Material['H20'].value = float(
                element.find('Parameter/[@Name = "FLC4"]/\
ParamValue[@Name = "Frequency"]').get('Value'))

            Material['H21'].value = float(
                element.find('Parameter/[@Name = "FLC4"]/\
ParamValue[@Name = "PulseLength"]').get('Value'))

        except AttributeError:
            pass

        #  Минимальный уровень мощности
        Material['G22'].value = float(
            element.find('Parameter/[@Name = "FLC"]/\
ParamValue[@Name = "MinPower"]').get('Value'))

        #  Тип регулирования
        Material['G23'].value = TypeOfRegulation[int(
            (element.find('Parameter/[@Name = "FLC"]/\
ParamValue[@Name = "RegMode"]').get('Value')))]

        #  Сопло
        Material['C33'].value = element.find(
            'Parameter/[@Name = "AMP_Jet"]/\
ParamValue[@Name = "Value"]').get('Value')

        #  Положение сопла
        Material['C9'].value = element.find(
            'Parameter/[@Name = "AMP_Focus"]/\
ParamValue[@Name = "Value"]').get('Value')

        #  Линза
        Material['C32'].value = element.find(
            'Parameter/[@Name = "AMP_Lens"]/\
ParamValue[@Name = "Value"]').get('Value')

#  Резка
        #  Скорость
        Material['D3'].value = float(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "Speed"]').get('Value'))

        #  Скорость FineCut
        Material['D4'].value = float(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "FineCutSpeed"]').get('Value'))

        #  Время прожига(указывается неявно)

        RepeatCount = int(
            float(element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "RepeatCount"]').get('Value')))

        BurnTime = int(
            float(element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "BurnTime"]').get('Value')))

        Wait = int(
            float(element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "WaitTime"]').get('Value')))

        FullPowerBurn = int(
            float(element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "FullPowerBurnTime"]').get('Value')))

        if RepeatCount > 1 and Wait != 0 and FullPowerBurn != 0\
                and BurnTime != 0:
            Material['D5'].value = (str(RepeatCount) + ':' + str(BurnTime) +
            ','+str(Wait)+','+str(FullPowerBurn))
        elif RepeatCount > 1 and Wait != 0 and FullPowerBurn == 0:
            Material['D5'].value = (str(RepeatCount) + ':' +
            str(BurnTime) + ',' + str(Wait))
        elif RepeatCount > 1 and Wait == 0 and FullPowerBurn == 0\
            and BurnTime != 0:
            Material['D5'].value = str(RepeatCount)+':'+str(BurnTime)
        elif RepeatCount == 1 and Wait == 0:
            Material['D5'].value = str(BurnTime)
        elif RepeatCount == 1 and Wait != 0 and FullPowerBurn != 0:
            Material['D5'].value = (str(BurnTime) + ',' +
            str(Wait) + ',' + str(FullPowerBurn))
        elif RepeatCount == 1 and Wait != 0 and FullPowerBurn == 0:
            Material['D5'].value = str(BurnTime)+','+str(Wait)

        #  Эквидистанта
        Material['D6'].value = float(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "EquidistantParameters"]/Value').get('Offset'))

        #  Скорость гравировки
        Material['C28'].value = float(
            element.find('Parameter/[@Name = "Tool: Гравировка"]/\
ParamValue[@Name = "Speed"]').get('Value'))

        #  Выпаривание(может быть не прописано для некоторых режимов)
        try:
            #  Высота фокуса
            Material['G40'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "FocusPosition"]').get('Value'))/10

            #  Высота подскока
            Material['G37'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "JumpHeight"]').get('Value'))/10

            #  Время подскока
            Material['G41'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "JumpTime"]').get('Value'))

            #  Положение линзы
            Material['G42'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "LensPosition"]').get('Value'))/10

            #  Подскок линзы
            Material['G39'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "LensJump"]').get('Value'))/10

            #  Высота гравировки
            Material['G43'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "EngraveHeight"]').get('Value'))/10

            #  Время опускания
            Material['G38'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "SlowDownTime"]').get('Value'))

            #  Без слежения
            Material['G35'].value = YesNot[int(
                int(element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "WithoutTracking"]').get('Value')))]

            #  Положение
            Material['H36'].value = int(
                element.find('Parameter/[@Name = "THC_THC_1"]/\
ParamValue[@Name = "ManualPosition"]').get('Value'))

            #  Скорость
            Material['C29'].value = float(
                element.find('Parameter/[@Name = "Tool: Лазер 1"]/\
ParamValue[@Name = "Speed"]').get('Value'))

            #  Режим выпаривания включен
            Material['C30'].value = YesNot[int(
                element.find('Parameter/[@Name = "Tool: Лазер 1"]/\
ParamValue[@Name = "ModeEnabled"]').get('Value'))]
        except AttributeError:
            pass

        #  Автоочистка сопла
        Material['C22'].value = YesNot[int(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "CleaningEnabled"]').get('Value'))]

        #  Количество врезок до очистки
        Material['C23'].value = float(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "PierceCountBeforeCleaning"]').get('Value'))

        #  Охлаждение при пробивке
        Material['C24'].value = YesNot[int(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "AirCoolingDuringPiercing"]').get('Value'))]

        #  Охлаждение при резке
        Material['C25'].value = YesNot[int(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "AirCoolingDuringCutting"]').get('Value'))]

        #  Автокалибровка сопла
        Material['C26'].value = YesNot[int(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "Autocalibration"]').get('Value'))]

        #  Предварительный прожиг
        Material['C27'].value = YesNot[int(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "PreBurning"]').get('Value'))]

        InputMetod = ['Без захода', "Заход по радиусу", "Заход по прямой"]

        OutputMetod = [
            'Без выхода', "Выход по радиусу",
            "Выход по прямой", 'Перемычка',
            'Выход по касательной без луча',
            "Выход за точку начала\
            обработки без луча"
            ]

        #  Способ входа
        Material['C12'].value = InputMetod[int(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "EquidistantParameters"]/Value').get('InputMethod'))]

        #  Длина захода
        Material['D14'].value = float(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "EquidistantParameters"]/Value').get('InputLength'))

        #  Радиус захода
        Material['D13'].value = float(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "EquidistantParameters"]/Value').get('InputRadius'))

        #  Способ выхода
        Material['C15'].value = OutputMetod[int(
            (element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "EquidistantParameters"]/Value').get('OutputMethod')))]

        #  Длина/Радиус выхода
        Material['D16'].value = float(
            element.find('Parameter/[@Name = "Tool: Лазер"]/\
ParamValue[@Name = "EquidistantParameters"]/Value').get('OutputLength'))

#  Микроперемычки(могут быть не указаны в режиме)

        try:
            Material['C18'].value = 'Да'
            #  Промежуток
            Material['D20'].value = float(
                element.find('Parameter/[@Name = "microjoints"]/\
ParamValue[@Name = "Period"]').get('Value'))

            #  Длина перемычки
            Material['D19'].value = float(
                element.find('Parameter/[@Name = "microjoints"]/\
ParamValue[@Name = "Length"]').get('Value'))

            #  Минимальная длина элемента
            Material['D21'].value = float(
                element.find('Parameter/[@Name = "microjoints"]/\
ParamValue[@Name = "Length"]').get('Value'))

        except AttributeError:
            Material['C18'].value = 'Нет'
            Material['D20'].value = 0
            Material['D19'].value = 0
            Material['D21'].value = 0

#  Газовая консоль

        #  Давление при резке
        try:
            Material['D38'].value = float(
                element.find('Parameter/[@Name = "GCB0"]/\
ParamValue[@Name = "Pressure"]').get('Value'))

        #  Газ при резке
            Material['D37'].value = TypeOfGas[int(
                element.find('Parameter/[@Name = "GCB0"]/\
ParamValue[@Name = "Input"]').get('Value'))]

        except AttributeError:
            pass
        try:
            #  Давление при выпаривании
            Material['D41'].value = float(
                element.find('Parameter/[@Name = "GCB0_1"]/\
ParamValue[@Name = "Pressure"]').get('Value'))

            #  Газ при выпаривании
            Material['D40'].value = TypeOfGas[int(
                element.find('Parameter/[@Name = "GCB0_1"]/\
ParamValue[@Name = "Input"]').get('Value'))]

        except AttributeError:
            pass

        #  Обратная гиперссылка на оглавление
        Material['A42'].hyperlink = "#Sheet!A1"
        Material['A42'].value = 'Назад к Оглавлению'
        Material['A42'].style = "Hyperlink"

        #  Создание словаря для последующей сортировки оглавления
        A.update({"#" + element.get('ID') + "!A1":TypeOfMaterial[int(Type)] +
                ' ' + str(float(element.get('Thickness'))) + ' мм. '
                        + element.get('Description')})
        i = i+1
#  Отсортированные гиперссылки оглавления
Result = sorted(A, key=A.get, reverse=False)
i = 1
for key in Result:
    Materials['Sheet'].cell(i, 1).hyperlink = key
    Materials['Sheet'].cell(i, 1).style = "Hyperlink"
    Materials['Sheet'].cell(i, 1).value = A[key]
    i = i+1

#  Сохранение итоговой таблицы
Materials.save('Materials'+' '+PowerType[int(Power)-1]+'.xlsx')
